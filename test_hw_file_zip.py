import os
import csv
import zipfile
from PyPDF2 import PdfReader
from openpyxl import load_workbook
import codecs

def test_create_zip():
    file_zip = zipfile.ZipFile('./resources/folder_zip.zip', 'w')
    file_zip.write('./resources/PDF_file.pdf')
    file_zip.write('./resources/XLSX_file.xlsx')
    file_zip.write('./resources/CSV_file.csv')


def test_for_file_pdf():
    with zipfile.ZipFile(os.path.abspath('./resources/folder_zip.zip')) as test_z_pdf:
        with test_z_pdf.open('resources/PDF_file.pdf') as test_for_pdf:
            test_for_pdf = PdfReader(test_for_pdf)
            text_check = test_for_pdf.pages[0].extract_text()
            assert ('QA ENGINEER' in text_check)


def test_for_file_xlsx():
    with zipfile.ZipFile(os.path.abspath('./resources/folder_zip.zip')) as test_z_xlsx:
        with test_z_xlsx.open('resources/XLSX_file.xlsx') as test_for_xlsx:
            test_for_xlsx = load_workbook(test_for_xlsx)
            sheet = test_for_xlsx.active
            data = sheet.cell(row=22, column=3).value
            assert data == 0.065


def test_for_file_csv():
    with zipfile.ZipFile(os.path.abspath('./resources/folder_zip.zip')) as test_z_csv:
        with test_z_csv.open('resources/CSV_file.csv') as test_for_csv:
            test_for_csv = csv.reader(codecs.iterdecode(test_for_csv, 'utf-8'))
            for line_number, line in enumerate(test_for_csv, 1):
                if line_number == 5:
                    assert line[5] == 'Pants'